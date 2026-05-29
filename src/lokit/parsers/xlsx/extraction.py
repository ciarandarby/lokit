from __future__ import annotations

import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import AsyncIterator, Iterator, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

from lokit.data.structure import Comment, Data, TranslationStatus

ExtractItem = tuple[str, Data]

_KNOWN_COLUMNS = frozenset({"id", "source", "target", "status", "comment"})


def _parse_base_lang(locale: str) -> str:
    return locale.replace("_", "-").split("-")[0].lower()


def _parse_status(value: str) -> TranslationStatus:
    normalized = value.strip().lower()
    try:
        return TranslationStatus(normalized)
    except ValueError:
        return TranslationStatus.UNKNOWN


def _cell_str(cell: Cell | MergedCell) -> str:
    if cell.value is None:
        return ""
    return str(cell.value)


def _infer_locales_from_filename(filepath: str) -> tuple[str, str | None]:
    stem = Path(filepath).stem
    if "-" in stem:
        parts = stem.split("-")
        if len(parts) == 2:
            return parts[0], parts[1]
        if len(parts) == 4:
            return f"{parts[0]}-{parts[1]}", f"{parts[2]}-{parts[3]}"
    if "_" in stem:
        parts = stem.split("_")
        if len(parts) == 2:
            return parts[0], parts[1]
        if len(parts) == 4:
            return f"{parts[0]}_{parts[1]}", f"{parts[2]}_{parts[3]}"
    return "", None


@dataclass(slots=True)
class _AsyncExtractionResult:
    item: Optional[ExtractItem] = None
    error: Optional[BaseException] = None
    done: bool = False


class AsyncXlsxExtraction:
    def __init__(self, extractor: XlsxExtractor) -> None:
        self._extractor = extractor
        self._queue: asyncio.Queue[_AsyncExtractionResult] = asyncio.Queue()
        self._producer: asyncio.Task[None] | None = None

    def __aiter__(self) -> AsyncXlsxExtraction:
        return self

    async def __anext__(self) -> ExtractItem:
        if self._producer is None:
            self._start()

        result = await self._queue.get()
        if result.done:
            await self._finish()
            raise StopAsyncIteration
        if result.error is not None:
            await self._finish()
            raise result.error
        if result.item is None:
            await self._finish()
            raise StopAsyncIteration
        return result.item

    def _start(self) -> None:
        loop = asyncio.get_running_loop()

        def produce() -> None:
            try:
                for item in self._extractor.extract():
                    loop.call_soon_threadsafe(
                        self._queue.put_nowait,
                        _AsyncExtractionResult(item=item),
                    )
            except BaseException as exc:
                loop.call_soon_threadsafe(
                    self._queue.put_nowait,
                    _AsyncExtractionResult(error=exc),
                )
            finally:
                loop.call_soon_threadsafe(
                    self._queue.put_nowait,
                    _AsyncExtractionResult(done=True),
                )

        self._producer = asyncio.create_task(asyncio.to_thread(produce))

    async def _finish(self) -> None:
        if self._producer is not None:
            await self._producer


class XlsxExtractor:
    def __init__(
        self,
        filepath: str,
        source_locale: str = "",
        target_locale: str | None = None,
    ) -> None:
        self.filepath: str = filepath

        if source_locale:
            self.source_locale: str = source_locale
            self.target_locale: str | None = target_locale
        else:
            inferred_source, inferred_target = _infer_locales_from_filename(filepath)
            self.source_locale = inferred_source
            self.target_locale = target_locale or inferred_target

        self.source_language: str | None = (
            _parse_base_lang(self.source_locale) if self.source_locale else None
        )
        self.target_language: str | None = (
            _parse_base_lang(self.target_locale) if self.target_locale else None
        )

        self.export_origin: str = ""
        self.export_timestamp: str = ""
        self.extensions: dict[str, str] = {"input_format": "xlsx"}

    def extract(self) -> Iterator[ExtractItem]:
        wb = load_workbook(self.filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return

            rows = ws.iter_rows()
            header_row = next(rows, None)
            if header_row is None:
                return

            headers: list[str] = [_cell_str(c).strip().lower() for c in header_row]
            col_map: dict[str, int] = {name: i for i, name in enumerate(headers) if name}
            has_id = "id" in col_map
            extra_columns = [h for h in headers if h and h not in _KNOWN_COLUMNS]

            for index, row in enumerate(rows):
                cells = list(row)

                def get(col: str) -> str:
                    idx = col_map.get(col)
                    if idx is None or idx >= len(cells):
                        return ""
                    return _cell_str(cells[idx])

                unit_id = get("id") if has_id else ""
                if not unit_id:
                    unit_id = f"xlsx:{index}"

                source = get("source")
                raw_target = get("target")
                target = raw_target if raw_target else None
                status = _parse_status(get("status")) if get("status") else TranslationStatus.UNKNOWN

                comments: list[Comment] = []
                comment_text = get("comment").strip()
                if comment_text:
                    comments.append(Comment(context=comment_text))

                extensions: dict[str, str] = {}
                for col in extra_columns:
                    val = get(col)
                    if val:
                        extensions[col] = val

                yield unit_id, Data(
                    source=source,
                    target=target,
                    status=status,
                    comments=comments,
                    extensions=extensions,
                )
        finally:
            wb.close()

    def extract_async(self) -> AsyncIterator[ExtractItem]:
        return AsyncXlsxExtraction(self)
